#!/usr/bin/env python3
"""
Export Script: ZKTeco Google Ads Campaign + GMC Product Feed
Generates: XLSX, CSV, and JSON files for both deliverables.
"""

import csv
import json
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUT_DIR = "h:/scraper/output"
os.makedirs(OUT_DIR, exist_ok=True)

# ──────────────────────────────────────────────
# STYLING HELPERS
# ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def style_data(ws, row_count, col_count):
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL

def auto_width(ws, col_count, max_width=50):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3


# ──────────────────────────────────────────────
# GMC FEED DATA
# ──────────────────────────────────────────────
GMC_HEADERS = [
    "id", "title", "description", "link", "image_link",
    "availability", "price", "brand", "condition", "gtin",
    "mpn", "identifier_exists", "google_product_category", "product_type"
]

GMC_ROWS = [
    ["ZC31","Zebra ZC300 Single-Sided Card Printer","Single-sided direct-to-card dye-sublimation printer. 300 dpi, USB 2.0 & Ethernet, 100-card hopper.","https://nahamtech.org/product/zebra-zc300-single","https://img.nahamtech.org/products/card-printers/zebra-zc31.jpg","in_stock","2700.00 SAR","Zebra","new","","zc31","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["ZC32","Zebra ZC300 Double-Sided Card Printer","Dual-sided direct-to-card dye-sublimation printer. 300 dpi, USB 2.0 & Ethernet, 100-card hopper.","https://nahamtech.org/product/zebra-zc300-double","https://img.nahamtech.org/products/card-printers/zebra-zc31.jpg","in_stock","3800.00 SAR","Zebra","new","","zc32","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["SMART-51S","SMART Single-Sided ID Card Printer","Single-sided smart card printer for ID cards and badges.","https://nahamtech.org/product/smart-51s","https://img.nahamtech.org/products/card-printers/smart-printer-1.png","in_stock","2800.00 SAR","IDP","new","","[653170] SMART-51S","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["SMART-51D","SMART Dual-Sided ID Card Printer","Dual-sided smart card printer for ID cards and badges.","https://nahamtech.org/product/smart-51d","https://img.nahamtech.org/products/card-printers/smart-printer-1.png","in_stock","4200.00 SAR","IDP","new","","[653171] SMART-51D","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["PM2-0001-M","Evolis Primacy 2 Single-Sided Card Printer","Single-sided card printer by Evolis for professional ID cards.","https://nahamtech.org/product/evolis-primacy2-single","https://img.nahamtech.org/products/card-printers/evolis-primacy-printer-1.jpg","out_of_stock","4800.00 SAR","Evolis","new","","PM2-0001-M","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["PM2-0025-M","Evolis Primacy 2 Dual-Sided Card Printer","Dual-sided card printer by Evolis for professional ID cards.","https://nahamtech.org/product/evolis-primacy2-dual","https://img.nahamtech.org/products/card-printers/evolis-primacy-printer-1.jpg","out_of_stock","6000.00 SAR","Evolis","new","","PM2-0025-M","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["Z72-000C0000EM00","Zebra ZXP Series 7 Dual-Sided Card Printer","Industrial dual-sided card printer with USB and Ethernet.","https://nahamtech.org/product/zebra-zxp7","https://img.nahamtech.org/products/card-printers/Zebra-ZXP-Series-7-Dual-Sided-Card-Printer.jpg","in_stock","7500.00 SAR","Zebra","new","","Z72-000C0000EM00","yes","Electronics > Print & Scan > Printers","Printers > Card Printers"],
    ["ACL003","Evolis ACL003 Cleaning Cards","Cleaning cards for Evolis card printers. Maintains print quality.","https://nahamtech.org/product/evolis-acl003","https://img.nahamtech.org/products/card-printers/Evolis-ACL003-Cleaning-Cards.jpg","in_stock","50.00 SAR","Evolis","new","","ACL003","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Cleaning Cards"],
    ["R5F208M100","Evolis Primacy 2 YMCKO Color Ribbon 200 prints","Color ribbon for Evolis Primacy 2 card printer, 200 prints.","https://nahamtech.org/product/evolis-primacy2-ribbon-color","https://img.nahamtech.org/products/card-printers/Evolis-Primacy-2-YMCKO-Color-Ribbon.jpg","in_stock","280.00 SAR","Evolis","new","","RCT223NAAA-R5F202M100-R5F208M100","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["659366","IDP SMART YMCKO Color Ribbon 250 prints","Color ribbon for IDP SMART card printers, 250 prints.","https://nahamtech.org/product/smart-ribbon-color","https://img.nahamtech.org/products/card-printers/IDP-Smart-YMCKO-Color-Ribbon-250-prints.jpg","in_stock","220.00 SAR","IDP","new","","659366","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["800077-740EM","Zebra ZXP7 YMCKO Color Ribbon 250 prints","Color ribbon for Zebra ZXP Series 7 card printer.","https://nahamtech.org/product/zebra-zxp7-ribbon-color","https://img.nahamtech.org/products/card-printers/Zebra-ZXP7-YMCKO-Color-Ribbon-250-prints.jpg","in_stock","250.00 SAR","Zebra","new","","800077-740EM_800077-701","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["800033-840","Zebra ZXP3 YMCKO Color Ribbon 200 prints","Color ribbon for Zebra ZXP3 card printer, 200 prints.","https://nahamtech.org/product/zebra-zxp3-ribbon","https://img.nahamtech.org/products/card-printers/Zebra-ZXP3-YMCKO-200-prints.jpg","in_stock","170.00 SAR","Zebra","new","","800033-840","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["800300-250EM","Zebra ZC100/ZC300 YMCKO Ribbon 200 prints","Color and monochrome ribbon for Zebra ZC100 and ZC300 series.","https://nahamtech.org/product/zebra-zc-ribbon-color","https://img.nahamtech.org/products/card-printers/Zebra-ZC100-ZC300-YMCKO-200-prints.jpg","in_stock","170.00 SAR","Zebra","new","","800300-250EM","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["800300-301EM","Zebra ZC100/ZC300 Black Monochrome Ribbon","Black monochrome ribbon for Zebra ZC100 and ZC300 series.","https://nahamtech.org/product/zebra-zc-ribbon-mono","https://img.nahamtech.org/products/card-printers/Zebra-ZC100-ZC300-YMCKO-200-prints-black.png","in_stock","160.00 SAR","Zebra","new","","800300-301EM","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["PVC-CARD","Blank White PVC Cards CR80 (250 pcs)","Pack of 250 blank white PVC cards, CR80 size, for card printers.","https://nahamtech.org/product/pvc-cards-250","https://img.nahamtech.org/products/accessories/Blank-White-PVC-Cards-CR80.jpg","in_stock","115.00 SAR","","new","","PVC-CARD","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Cards"],
    ["S50-S70","Mifare RFID Smart Cards 1K S50 / 4K S70 (10 pcs)","Pack of 10 Mifare RFID contactless smart cards.","https://nahamtech.org/product/mifare-rfid-cards","https://img.nahamtech.org/products/accessories/Mifare-RFID-Cards-1K-S50-4K-S70.jpg","in_stock","40.00 SAR","","new","","S50-S70","yes","Electronics > Access Control","Accessories > NFC Cards"],
    ["104523-111","Zebra Blank White PVC Card 30 mil","Zebra-branded blank white PVC cards, 30 mil thickness.","https://nahamtech.org/product/zebra-pvc-card","https://img.nahamtech.org/products/accessories/Blank-White-PVC-Cards-CR80.jpg","in_stock","60.00 SAR","Zebra","new","","104523-111","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Cards"],
    ["TM-TAG03-4100","ZKTeco RFID EM Keyfob 125KHz TK4100 (10 pcs)","Pack of 10 RFID EM keyfobs, 125KHz TK4100, for access control.","https://nahamtech.org/product/zkteco-keyfob-tk4100","https://img.nahamtech.org/products/accessories/TK4100-Proximity-ID-Cards-and-Keyfobs.jpg","in_stock","60.00 SAR","ZKTeco","new","","TM-TAG03-4100","yes","Electronics > Access Control","Accessories > NFC Cards"],
    ["105999-701","Zebra ZXP Series 7 Cleaning Kit","Cleaning kit for Zebra ZXP Series 7 card printer.","https://nahamtech.org/product/zebra-zxp7-cleaning","https://img.nahamtech.org/products/card-printers/Zebra-ZXP-Series-7-Cleaning-Kit.jpg","in_stock","100.00 SAR","Zebra","new","","105999-701","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Cleaning Cards"],
    ["800077-701","Zebra ix Series Monochrome Ribbon for ZXP 7","Black monochrome ribbon for Zebra ZXP Series 7.","https://nahamtech.org/product/zebra-zxp7-mono-ribbon","https://img.nahamtech.org/products/card-printers/Zebra-ix-Series-monochrome-ribbon-for-ZXP-Series-7.jpg","in_stock","180.00 SAR","Zebra","new","","800077-701","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Ink & Ribbons"],
    ["105999-311","Zebra ZC Series Cleaning Cards","Cleaning cards for Zebra ZC series card printers.","https://nahamtech.org/product/zebra-zc-cleaning","https://img.nahamtech.org/products/card-printers/Zebra-ZC-Series-Cleaning-Cards.jpg","in_stock","100.00 SAR","Zebra","new","","105999-311","yes","Electronics > Print & Scan > Printer Accessories","Accessories > Cleaning Cards"],
    ["TK4100","TK4100 Proximity ID Thin Card Blank (10 pcs)","Pack of 10 TK4100 proximity thin cards, blank, no serial.","https://nahamtech.org/product/tk4100-thin-blank","https://img.nahamtech.org/products/accessories/TK4100-Proximity-ID-Thin-Card-Blank-without-serial.jpeg","in_stock","20.00 SAR","ZKTeco","new","","TK4100","yes","Electronics > Access Control","Accessories > NFC Cards"],
    ["TK4100LN","TK4100 Proximity ID Thin Card with Long Serial (10 pcs)","Pack of 10 TK4100 proximity thin cards with long serial number.","https://nahamtech.org/product/tk4100-long-serial","https://img.nahamtech.org/products/accessories/blank-white-card-long-serial.jpg","in_stock","20.00 SAR","ZKTeco","new","","TK4100LN","yes","Electronics > Access Control","Accessories > NFC Cards"],
    ["TK4100SN","TK4100 Proximity ID Thin Card with Short Serial (10 pcs)","Pack of 10 TK4100 proximity thin cards with short serial number.","https://nahamtech.org/product/tk4100-short-serial","https://img.nahamtech.org/products/accessories/blank-white-card-short-serial.jpg","in_stock","20.00 SAR","ZKTeco","new","","TK4100SN","yes","Electronics > Access Control","Accessories > NFC Cards"],
    ["QW2120-BKK1S","Datalogic QuickScan Lite QW2100 1D Barcode Scanner","Wired 1D barcode scanner by Datalogic.","https://nahamtech.org/product/datalogic-qw2100","https://img.nahamtech.org/products/barcode-scanners/data-logic-barcode-scanner-1.jpg","out_of_stock","180.00 SAR","Datalogic","new","","QW2120-BKK1S","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["QW2420-BKK1S","Datalogic QuickScan Lite QW2420 2D Barcode Scanner","Wired 2D barcode scanner by Datalogic.","https://nahamtech.org/product/datalogic-qw2420","https://img.nahamtech.org/products/barcode-scanners/data-logic-barcode-scanner-1.jpg","out_of_stock","300.00 SAR","Datalogic","new","","QW2420-BKK1S","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["QW2520-BKK1S","Datalogic QuickScan QW2500 2D Barcode Scanner","Wired 2D barcode scanner by Datalogic.","https://nahamtech.org/product/datalogic-qw2500","https://img.nahamtech.org/products/barcode-scanners/data-logic-barcode-scanner-1.jpg","in_stock","350.00 SAR","Datalogic","new","","QW2520-BKK1S","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["MK7120-31A38","Honeywell Orbit 7120 Omnidirectional Barcode Scanner","Omnidirectional hands-free 1D barcode scanner by Honeywell.","https://nahamtech.org/product/honeywell-orbit-7120","https://img.nahamtech.org/products/barcode-scanners/honeywell-barcode-scanner-MK7120-31A38.jpg","in_stock","337.00 SAR","Honeywell","new","","MK7120-31A38","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["1470G2D-2USB-1-R","Honeywell Voyager XP 1470g 2D Barcode Scanner","Wired 2D barcode scanner by Honeywell.","https://nahamtech.org/product/honeywell-1470g","https://img.nahamtech.org/products/barcode-scanners/honewell1470g-barcode-scanner-1.jpg","in_stock","320.00 SAR","Honeywell","new","","1470G2D-2USB-1-R","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["1250G-2USB-1","Honeywell 1250g Laser 1D Barcode Scanner","Wired laser 1D barcode scanner by Honeywell.","https://nahamtech.org/product/honeywell-1250g","https://img.nahamtech.org/products/barcode-scanners/honeywell-1250g.jpg","in_stock","225.00 SAR","Honeywell","new","","1250G-2USB-1","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["1472G2D-2USB-5-R","Honeywell Voyager XP 1472g Wireless 2D Barcode Scanner","Wireless 2D barcode scanner by Honeywell with Bluetooth.","https://nahamtech.org/product/honeywell-1472g","https://img.nahamtech.org/products/barcode-scanners/honeywell-1472g.jpg","in_stock","650.00 SAR","Honeywell","new","","1472G2D-2USB-5-R","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["PRO-V6258","PRO-V 6258 2D Wired Barcode Scanner","Wired 2D barcode scanner by PRO-V.","https://nahamtech.org/product/prov-6258","https://img.nahamtech.org/products/barcode-scanners/PRO-V-6258.jpg","in_stock","92.00 SAR","PRO-V","new","","PRO-V6258","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["PRO-V6221RB","PRO-V 6221RB Wireless Bluetooth Barcode Scanner","Wireless Bluetooth barcode scanner by PRO-V.","https://nahamtech.org/product/prov-6221rb","https://img.nahamtech.org/products/barcode-scanners/PRO-V-6221RB_1.jpg","in_stock","200.00 SAR","PRO-V","new","","PRO-V6221RB","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["HDD-518W","Price Checker HDD-518W Windows","Windows-based price checker kiosk device.","https://nahamtech.org/product/price-checker-518w","https://img.nahamtech.org/products/barcode-scanners/master-price-checker.png","in_stock","1500.00 SAR","","new","","HDD-518W","yes","Electronics > POS Equipment","POS > Barcode Scanners"],
    ["PRO-V6266MBT","PRO-V 6266MBT Wireless 2D Barcode Scanner","Wireless 2D barcode scanner with Bluetooth.","https://nahamtech.org/product/prov-6266mbt","https://img.nahamtech.org/products/barcode-scanners/PRO-V6266MBT.jpg","in_stock","340.00 SAR","PRO-V","new","","PRO-V6266MBT","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["UNS-5320R","Uniscan UNS-5320R Wireless 2D Barcode Scanner","Wireless 2D barcode scanner by Uniscan.","https://nahamtech.org/product/uniscan-5320r","https://img.nahamtech.org/products/barcode-scanners/UniScan-UNS-5320R.jpg","out_of_stock","350.00 SAR","Uniscan","new","","UNS-5320R","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["LS2208","Zebra LS2208 Laser 1D Barcode Scanner","Wired laser 1D barcode scanner by Zebra.","https://nahamtech.org/product/zebra-ls2208","https://img.nahamtech.org/products/barcode-scanners/Zebra-LI2208_1-LS2208_1.jpg","in_stock","210.00 SAR","Zebra","new","","LS2208-SR20007R-UR","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["PRO-V8602H","PRO-V 8602H Desktop 2D Barcode Scanner","Desktop presentation 2D barcode scanner.","https://nahamtech.org/product/prov-8602h","https://img.nahamtech.org/products/barcode-scanners/PRO-V-8602H.png","in_stock","218.00 SAR","PRO-V","new","","PRO-V8602H","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["LI2208","Zebra LI2208 Linear 1D Barcode Scanner","Wired linear 1D barcode scanner by Zebra.","https://nahamtech.org/product/zebra-li2208","https://img.nahamtech.org/products/barcode-scanners/Zebra-LI2208_1-LS2208_1.jpg","out_of_stock","180.00 SAR","Zebra","new","","LI2208-SR7U2100SGW","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["CBA-U01-S07ZAR","Zebra USB Cable CBA-U01-S07ZAR","USB cable for Zebra barcode scanners.","https://nahamtech.org/product/zebra-usb-cable","","in_stock","150.00 SAR","Zebra","new","","CBA-U01-S07ZAR","yes","Electronics > Cables","Accessories"],
    ["DS9308","Zebra DS9308 Desktop 2D Barcode Scanner","Desktop presentation 2D barcode scanner by Zebra.","https://nahamtech.org/product/zebra-ds9308","https://img.nahamtech.org/products/barcode-scanners/zebra-ds9308.jpg","in_stock","400.00 SAR","Zebra","new","","DS9308-SR4U2100AZE","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["DS2208","Zebra DS2208 2D Barcode Scanner (with Stand)","Wired 2D barcode scanner by Zebra, includes stand.","https://nahamtech.org/product/zebra-ds2208","https://img.nahamtech.org/products/barcode-scanners/Zebra-DS2208_1.jpg","in_stock","350.00 SAR","Zebra","new","","DS2208-SR7U2100SGW","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["DS2278","Zebra DS2278 Wireless 2D Barcode Scanner","Wireless 2D barcode scanner by Zebra.","https://nahamtech.org/product/zebra-ds2278","https://img.nahamtech.org/products/barcode-scanners/zebra-ds2278_1.jpg","in_stock","670.00 SAR","Zebra","new","","DS2278-SR7U2100PRW","yes","Electronics > Barcode Scanners","POS > Barcode Scanners"],
    ["AL-280PL","ZKTeco AL-280PL Mounting Bracket (Interior Doors)","L-shaped mounting bracket for ZKTeco magnetic locks.","https://nahamtech.org/product/zkteco-al280pl","https://img.nahamtech.org/products/ZKTeco/AL-280PZ_AL-280PL.jpg","in_stock","30.00 SAR","ZKTeco","new","","AL-280PL","yes","Electronics > Access Control","Security > Access Control"],
    ["AL-280PU","ZKTeco AL-280PU Mounting Bracket (Glass Doors)","U-shaped mounting bracket for glass door installation.","https://nahamtech.org/product/zkteco-al280pu","https://img.nahamtech.org/products/ZKTeco/AL-280PU.jpg","in_stock","35.00 SAR","ZKTeco","new","","AL-280PU","yes","Electronics > Access Control","Security > Access Control"],
    ["AL-280PZ","ZKTeco AL-280PZ Mounting Bracket (Z-Shape)","Z-shaped mounting bracket for ZKTeco magnetic locks.","https://nahamtech.org/product/zkteco-al280pz","https://img.nahamtech.org/products/ZKTeco/AL-280PZ_AL-280PL.jpg","in_stock","40.00 SAR","ZKTeco","new","","AL-280PZ","yes","Electronics > Access Control","Security > Access Control"],
    ["FR1200","ZKTeco FR1200 Slave Fingerprint Reader","External fingerprint and RFID reader for access control panels.","https://nahamtech.org/product/zkteco-fr1200","https://img.nahamtech.org/products/ZKTeco/fr1200.png","in_stock","270.00 SAR","ZKTeco","new","","FR1200","yes","Electronics > Access Control","Security > Access Control"],
    ["GL300","ZKTeco GL300 Smart Glass Door Lock","Smart lock for frameless glass doors with fingerprint, RFID and PIN.","https://nahamtech.org/product/zkteco-gl300","https://img.nahamtech.org/products/ZKTeco/GL300.jpeg","in_stock","550.00 SAR","ZKTeco","new","","GL300","yes","Electronics > Access Control > Door Locks","Security > Access Control"],
    ["UA760","ZKTeco UA760 Time Attendance (Finger + Card + WiFi)","Time attendance terminal with fingerprint, RFID card, PIN and WiFi.","https://nahamtech.org/product/zkteco-ua760","https://img.nahamtech.org/products/ZKTeco/UA760.jpeg","in_stock","360.00 SAR","ZKTeco","new","","UA760","yes","Electronics > Access Control","Time Attendance > Fingerprint Devices"],
    ["MB10","ZKTeco MB10 Time Attendance (Face + Fingerprint)","Compact multi-biometric terminal with face and fingerprint recognition.","https://nahamtech.org/product/zkteco-mb10","https://img.nahamtech.org/products/ZKTeco/MB10.png","in_stock","250.00 SAR","ZKTeco","new","","MB10","yes","Electronics > Access Control","Time Attendance > Fingerprint Devices"],
    ["UFACE800","ZKTeco UFace800 Time Attendance (Face + Finger + Card + PIN)","Multi-biometric terminal with 8-inch touchscreen. Face, fingerprint, card and PIN.","https://nahamtech.org/product/zkteco-uface800","https://img.nahamtech.org/products/ZKTeco/uFace-800-1.jpg","in_stock","810.00 SAR","ZKTeco","new","","UFACE800","yes","Electronics > Access Control","Time Attendance > Face Recognition"],
    ["F18","ZKTeco F18 Access Control (Fingerprint + Card + PIN)","Biometric access control device with fingerprint, RFID card and password.","https://nahamtech.org/product/zkteco-f18","https://img.nahamtech.org/products/ZKTeco/F18_1.jpg","in_stock","450.00 SAR","ZKTeco","new","","F18","yes","Electronics > Access Control","Security > Access Control"],
    ["UFACE-METAL-BOX","ZKTeco UFace Metal Box","Protective metal enclosure for ZKTeco UFace series devices.","https://nahamtech.org/product/zkteco-uface-metalbox","https://img.nahamtech.org/products/ZKTeco/uface-metal-box.png","in_stock","250.00 SAR","ZKTeco","new","","UFACE METAL BOX","yes","Electronics > Access Control","Accessories"],
    ["BATTERY-UFACE","ZKTeco UFace Backup Battery","Backup battery for ZKTeco UFace series devices.","https://nahamtech.org/product/zkteco-uface-battery","https://img.nahamtech.org/products/ZKTeco/BATTERY-UFACE.jpg","out_of_stock","120.00 SAR","ZKTeco","new","","BATTERY-UFACE","yes","Electronics > Access Control","Accessories"],
    ["TLEB101","ZKTeco TLEB101 Touchless Exit Button","Infrared touchless exit button for access control.","https://nahamtech.org/product/zkteco-tleb101","https://img.nahamtech.org/products/ZKTeco/TLEB101.jpg","in_stock","50.00 SAR","ZKTeco","new","","TLEB101","yes","Electronics > Access Control","Security > Access Control"],
    ["ZK4500","ZKTeco ZK4500 USB Fingerprint Scanner","USB fingerprint reader for enrollment and verification.","https://nahamtech.org/product/zkteco-zk4500","https://img.nahamtech.org/products/ZKTeco/ZK4500.png","out_of_stock","250.00 SAR","ZKTeco","new","","ZK4500","yes","Electronics > Access Control","Time Attendance > Fingerprint Devices"],
    ["SPEEDFACE-V5L-WIFI","ZKTeco SpeedFace V5L WiFi (Face + Palm + PIN)","Advanced biometric terminal with face, palm vein and PIN recognition. WiFi enabled.","https://nahamtech.org/product/zkteco-speedface-v5l-wifi","https://img.nahamtech.org/products/ZKTeco/SPEEDFACE-V5L-WIFI.png","in_stock","950.00 SAR","ZKTeco","new","","SPEEDFACE-V5L-WIFI","yes","Electronics > Access Control","Time Attendance > Face Recognition"],
    ["MB2000","ZKTeco MB2000 Time Attendance (Face + Fingerprint)","Multi-biometric terminal with face and fingerprint recognition.","https://nahamtech.org/product/zkteco-mb2000","https://img.nahamtech.org/products/ZKTeco/MB2000_1.png","in_stock","450.00 SAR","ZKTeco","new","","MB2000","yes","Electronics > Access Control","Time Attendance > Face Recognition"],
    ["SPEEDFACE-M4","ZKTeco SpeedFace M4 (Face + Palm + PIN)","Biometric terminal with face and palm recognition.","https://nahamtech.org/product/zkteco-speedface-m4","https://img.nahamtech.org/products/ZKTeco/SPEEDFACE-M4.jpg","in_stock","900.00 SAR","ZKTeco","new","","SPEEDFACE-M4","yes","Electronics > Access Control","Time Attendance > Face Recognition"],
    ["SPEEDFACE-M4-WIFI","ZKTeco SpeedFace M4 WiFi (Face + Palm + PIN)","Biometric terminal with face and palm recognition. WiFi enabled.","https://nahamtech.org/product/zkteco-speedface-m4-wifi","https://img.nahamtech.org/products/ZKTeco/SPEEDFACE-M4.jpg","in_stock","1100.00 SAR","ZKTeco","new","","SPEEDFACE-M4/WIFI","yes","Electronics > Access Control","Time Attendance > Face Recognition"],
    ["SLK20R","ZKTeco SLK20R USB Fingerprint Reader","USB live fingerprint scanner with silicone sensor.","https://nahamtech.org/product/zkteco-slk20r","https://img.nahamtech.org/products/ZKTeco/SLK20R.jpg","out_of_stock","240.00 SAR","ZKTeco","new","","SLK20R","yes","Electronics > Access Control","Time Attendance > Fingerprint Devices"],
    ["U650-C","ZKTeco U650-C Time Attendance","Time attendance terminal with fingerprint and card.","https://nahamtech.org/product/zkteco-u650c","https://img.nahamtech.org/products/ZKTeco/u-650C.png","out_of_stock","500.00 SAR","ZKTeco","new","","U650-C","yes","Electronics > Access Control","Time Attendance > Fingerprint Devices"],
    ["F22","ZKTeco F22 WiFi Access Control (Fingerprint + Card + PIN)","Biometric access control with WiFi, fingerprint, RFID card and password.","https://nahamtech.org/product/zkteco-f22","https://img.nahamtech.org/products/ZKTeco/F22_1.png","in_stock","330.00 SAR","ZKTeco","new","","F22","yes","Electronics > Access Control","Security > Access Control"],
    ["UFACE800-WIFI","ZKTeco UFace800 WiFi (Face + Finger + Card + PIN)","Multi-biometric terminal with WiFi. Face, fingerprint, card and PIN.","https://nahamtech.org/product/zkteco-uface800-wifi","https://img.nahamtech.org/products/ZKTeco/uFace-800-1.jpg","in_stock","850.00 SAR","ZKTeco","new","","UFACE800/WIFI","yes","Electronics > Access Control","Time Attendance > Face Recognition"],
    ["ZK9500","ZKTeco ZK9500 USB Fingerprint Scanner","USB fingerprint reader with optical sensor.","https://nahamtech.org/product/zkteco-zk9500","https://img.nahamtech.org/products/ZKTeco/ZK9500.jpg","in_stock","200.00 SAR","ZKTeco","new","","ZK9500","yes","Electronics > Access Control","Time Attendance > Fingerprint Devices"],
    ["AL-280-LED","ZKTeco AL-280 LED Magnetic Lock (270 kg)","Electromagnetic lock with LED indicator, 270 kg holding force.","https://nahamtech.org/product/zkteco-al280-led","https://img.nahamtech.org/products/ZKTeco/AL-280_LED.jpeg","in_stock","90.00 SAR","ZKTeco","new","","AL-280 LED","yes","Electronics > Access Control","Security > Access Control"],
    ["X7","ZKTeco X7 Access Control (Fingerprint + Card + PIN)","Compact access control device with fingerprint, card and PIN.","https://nahamtech.org/product/zkteco-x7","https://img.nahamtech.org/products/ZKTeco/x7.png","in_stock","170.00 SAR","ZKTeco","new","","X7","yes","Electronics > Access Control","Security > Access Control"],
    ["PRO-V411U","PRO-V411U Barcode Printer (USB)","Thermal/thermal transfer barcode label printer with USB.","https://nahamtech.org/product/prov-411u","https://img.nahamtech.org/products/barcode-printers/PRO-V411BU.jpg","in_stock","350.00 SAR","PRO-V","new","","PRO-V411U","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["PRO-V411BU","PRO-V411BU Barcode Printer (USB + Bluetooth)","Thermal/thermal transfer barcode label printer with USB and Bluetooth.","https://nahamtech.org/product/prov-411bu","https://img.nahamtech.org/products/barcode-printers/PRO-V411BU.jpg","in_stock","450.00 SAR","PRO-V","new","","PRO-V411BU","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["ZD22042-T0EG00EZ","Zebra ZD220 Barcode Printer (Thermal Transfer)","Desktop barcode label printer by Zebra.","https://nahamtech.org/product/zebra-zd220","https://img.nahamtech.org/products/barcode-printers/zebra-ZD22042-T0EG00EZ.jpg","in_stock","650.00 SAR","Zebra","new","","ZD22042-T0EG00EZ","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["ZT41142-T0E0000Z","Zebra ZT411 Industrial Barcode Printer","Industrial barcode label printer by Zebra, high-volume.","https://nahamtech.org/product/zebra-zt411","https://img.nahamtech.org/products/barcode-printers/ZT41142-T0E0000Z_1.jpg","in_stock","3200.00 SAR","Zebra","new","","ZT41142-T0E0000Z","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["ZD23042-30EC00EZ","Zebra ZD230 Barcode Printer (Network)","Desktop barcode label printer by Zebra with Ethernet.","https://nahamtech.org/product/zebra-zd230","https://img.nahamtech.org/products/barcode-printers/zebra-ZD23042-30EC00EZ.jpg","in_stock","850.00 SAR","Zebra","new","","ZD23042-30EC00EZ","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["GS-2408U","Gainscha GS-2408U Barcode Printer (USB)","Desktop barcode label printer by Gainscha with USB.","https://nahamtech.org/product/gainscha-gs2408u","https://img.nahamtech.org/products/barcode-printers/GS-2408U_GS-2408DC.png","in_stock","350.00 SAR","Gainscha","new","","GS-2408U","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["GS-2408DC","Gainscha GS-2408DC Barcode Printer (Bluetooth)","Desktop barcode label printer by Gainscha with Bluetooth.","https://nahamtech.org/product/gainscha-gs2408dc","https://img.nahamtech.org/products/barcode-printers/GS-2408U_GS-2408DC.png","in_stock","550.00 SAR","Gainscha","new","","GS-2408DC","yes","Electronics > Print & Scan > Printers","Printers > Barcode Printers"],
    ["EDA52","Honeywell EDA52 Handheld Mobile Computer","Android 11, 64GB/4GB, 13MP camera, rugged handheld scanner.","https://nahamtech.org/product/honeywell-eda52","https://img.nahamtech.org/products/inventory-and-money-counters/honeywell-inventory.jpg","in_stock","1330.00 SAR","Honeywell","new","","EDA52-11AE64N21RK","yes","Electronics > Barcode Scanners","Inventory > Handheld Computers"],
    ["TC15","Zebra TC15 Handheld Mobile Computer","Android rugged handheld, 64GB/4GB, 5G capable.","https://nahamtech.org/product/zebra-tc15","https://img.nahamtech.org/products/inventory-and-money-counters/Zebra-TC15.jpg","in_stock","1400.00 SAR","Zebra","new","","TC15BK-1PE14S-A6","yes","Electronics > Barcode Scanners","Inventory > Handheld Computers"],
    ["ST-618+","Master ST-618+ Touchscreen POS Terminal","15-inch touchscreen POS terminal, Intel i5 4th gen.","https://nahamtech.org/product/master-st618plus","https://img.nahamtech.org/products/cashier/master-1.jpg","in_stock","1100.00 SAR","","new","","ST-618+","yes","Electronics > POS Equipment","POS > Cashier Systems"],
    ["APEXA-CORE","POSBANK APEXA Core Touchscreen POS Terminal","15-inch POS terminal, Intel i5 11th gen, grey.","https://nahamtech.org/product/posbank-apexa-core","https://img.nahamtech.org/products/cashier/APEXA-TL-1.jpg","in_stock","3500.00 SAR","POSBANK","new","","APEXA CORE","yes","Electronics > POS Equipment","POS > Cashier Systems"],
    ["M20","Landi M20 Mobile Payment Terminal","4G handheld payment terminal with built-in printer, 32GB.","https://nahamtech.org/product/landi-m20","https://img.nahamtech.org/products/cashier/landi-m20-1.jpg","in_stock","750.00 SAR","Landi","new","","VN02000003 M20","yes","Electronics > POS Equipment","POS > Mobile Terminals"],
    ["TM-M30II-112A0","Epson TM-M30II Thermal Receipt Printer (Bluetooth)","80mm thermal receipt printer with USB, Ethernet and Bluetooth.","https://nahamtech.org/product/epson-tmm30ii-bt","https://img.nahamtech.org/products/cashier/Epson-TM-M30II-112A0_TM-M30II-122A0.jpg","in_stock","1100.00 SAR","Epson","new","","TM-M30II (112A0)","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["UP301B","UPOS UP301B Portable 3-inch Thermal Printer (Bluetooth)","Portable thermal receipt printer, 3-inch, USB and Bluetooth.","https://nahamtech.org/product/upos-up301b","https://img.nahamtech.org/products/cashier/UP301B.jpg","in_stock","250.00 SAR","UPOS","new","","UP301B","yes","Electronics > Print & Scan > Printers","Printers > Mobile Printers"],
    ["UP401B","UPOS UP401B Portable 4-inch Thermal Printer (Bluetooth)","Portable thermal receipt/label printer, 4-inch, USB and Bluetooth.","https://nahamtech.org/product/upos-up401b","https://img.nahamtech.org/products/cashier/UP401B.jpg","in_stock","450.00 SAR","UPOS","new","","UP401B","yes","Electronics > Print & Scan > Printers","Printers > Mobile Printers"],
    ["TQ30","ENOTEQ TQ30 Electronic Barcode Label Scale 30kg","Label printing electronic scale, 30kg capacity, built-in barcode.","https://nahamtech.org/product/enoteq-tq30","https://img.nahamtech.org/products/cashier/ENOTEQ.jpg","in_stock","750.00 SAR","ENOTEQ","new","","TQ30","yes","Electronics > POS Equipment","POS > Scales"],
    ["DS-MDT201","Hikvision DS-MDT201 Handheld Mobile Computer","5.5-inch handheld, 64GB/4GB, Android, barcode scanner.","https://nahamtech.org/product/hikvision-mdt201","https://img.nahamtech.org/products/inventory-and-money-counters/Hikvision-DS-MDT201.jpg","in_stock","868.60 SAR","Hikvision","new","","DS-MDT201","yes","Electronics > Barcode Scanners","Inventory > Handheld Computers"],
    ["DS-MDT202","Hikvision DS-MDT202 Handheld Mobile Computer","Android 14 handheld terminal with barcode reader and camera.","https://nahamtech.org/product/hikvision-mdt202","https://img.nahamtech.org/products/inventory-and-money-counters/Hikvision-DS-MDT202.jpg","out_of_stock","1100.00 SAR","Hikvision","new","","DS-MDT202","yes","Electronics > Barcode Scanners","Inventory > Handheld Computers"],
    ["ST-4000","SEMTOM ST-4000 Money Counter & Sorter (Dual CIS)","Dual CIS money counter and sorter with two pockets.","https://nahamtech.org/product/semtom-st4000","https://img.nahamtech.org/products/inventory-and-money-counters/SEMTOM-ST-4000.png","out_of_stock","3695.00 SAR","SEMTOM","new","","ST-4000","yes","Electronics > Cash Handling","Inventory > Money Counters"],
    ["PRO-V2000","PRO-V2000 Money Counter (UV/MG Counterfeit Detection)","Automatic money counter with UV and MG counterfeit detection.","https://nahamtech.org/product/prov-2000","https://img.nahamtech.org/products/inventory-and-money-counters/pro-v-2000_1.jpeg","in_stock","500.00 SAR","PRO-V","new","","PRO-V2000","yes","Electronics > Cash Handling","Inventory > Money Counters"],
    ["PRO-V420","PRO-V 420 Cash Drawer (Medium, 5 slots)","Medium-sized POS cash drawer with 5 bill compartments.","https://nahamtech.org/product/prov-420","https://img.nahamtech.org/products/cashier/pro-v-cash-drawer_1.jpg","in_stock","160.00 SAR","PRO-V","new","","PRO-V420","yes","Electronics > POS Equipment","POS > Cash Drawers"],
    ["PRO-V4042","PRO-V 4042 Cash Drawer (Medium, 5 slots)","Medium-sized POS cash drawer with 5 bill compartments.","https://nahamtech.org/product/prov-4042","https://img.nahamtech.org/products/cashier/pro-v-cash-drawer_1.jpg","in_stock","120.00 SAR","PRO-V","new","","PRO-V4042","yes","Electronics > POS Equipment","POS > Cash Drawers"],
    ["PRO-V460","PRO-V 460 Cash Drawer (Large, 5 slots)","Large POS cash drawer with 5 bill compartments.","https://nahamtech.org/product/prov-460","https://img.nahamtech.org/products/cashier/pro-v-cash-drawer_1.jpg","in_stock","220.00 SAR","PRO-V","new","","PRO-V460","yes","Electronics > POS Equipment","POS > Cash Drawers"],
    ["APEXA-X","POSBANK APEXA X Customer Display 9.7 inch","9.7-inch customer-facing display for POSBANK APEXA POS.","https://nahamtech.org/product/posbank-apexa-x-display","https://img.nahamtech.org/products/cashier/APEXA-TL-Customer-Monitor.jpg","in_stock","1000.00 SAR","POSBANK","new","","APEXA X - 9.7\"","yes","Electronics > POS Equipment","POS > Cashier Systems"],
    ["P06060049","SUNMI V2s Handheld POS with Label Printer (2GB/16GB)","Handheld POS terminal with label printer, 4G and NFC.","https://nahamtech.org/product/sunmi-v2s-label","https://img.nahamtech.org/products/cashier/SUNMI-V2s.jpg","in_stock","740.00 SAR","SUNMI","new","","P06060049","yes","Electronics > POS Equipment","POS > Mobile Terminals"],
    ["OT-WL06-327","Epson OT-WL06 WiFi Dongle Dual Band","Wireless LAN USB adapter for Epson printers.","https://nahamtech.org/product/epson-ot-wl06","https://img.nahamtech.org/products/cashier/Epson-OT-WL06-Wifi-Dongle.jpg","in_stock","350.00 SAR","Epson","new","","OT-WL06-327","yes","Electronics > Networking > Wireless Adapters","Accessories"],
    ["GA-E200IW","Gainscha GA-E200IW Thermal Receipt Printer (WiFi)","80mm thermal receipt printer with USB, Ethernet and WiFi.","https://nahamtech.org/product/gainscha-ga-e200iw","https://img.nahamtech.org/products/cashier/Gainscha-GA-E200IB_GA-E200IW.jpeg","in_stock","350.00 SAR","Gainscha","new","","GA-E200IW","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["GA-E200IB","Gainscha GA-E200IB Thermal Receipt Printer (Bluetooth)","80mm thermal receipt printer with USB, Ethernet and Bluetooth.","https://nahamtech.org/product/gainscha-ga-e200ib","https://img.nahamtech.org/products/cashier/Gainscha-GA-E200IB_GA-E200IW.jpeg","in_stock","350.00 SAR","Gainscha","new","","GA-E200IB","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["TM-M30II-122A0","Epson TM-M30II Thermal Receipt Printer (Ethernet)","80mm thermal receipt printer with USB and Ethernet.","https://nahamtech.org/product/epson-tmm30ii-eth","https://img.nahamtech.org/products/cashier/Epson-TM-M30II-112A0_TM-M30II-122A0.jpg","in_stock","850.00 SAR","Epson","new","","TM-M30II (122A0)","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["TM-T20III-011A0","Epson TM-T20III Thermal Receipt Printer (USB)","80mm thermal receipt printer with USB.","https://nahamtech.org/product/epson-tmt20iii-usb","https://img.nahamtech.org/products/cashier/Epson-TM-T20III-011A0-012A0.jpg","in_stock","500.00 SAR","Epson","new","","TM-T20III (011A0)","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["TM-T20III-012A0","Epson TM-T20III Thermal Receipt Printer (USB + Ethernet)","80mm thermal receipt printer with USB and Ethernet.","https://nahamtech.org/product/epson-tmt20iii-eth","https://img.nahamtech.org/products/cashier/Epson-TM-T20III-011A0-012A0.jpg","in_stock","670.00 SAR","Epson","new","","TM-T20III (012A0)","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["PRO-V326U","PRO-V326U Thermal Receipt Printer 80mm","Compact 80mm thermal receipt printer.","https://nahamtech.org/product/prov-326u","https://img.nahamtech.org/products/cashier/PRO-V326U-80mm.jpg","in_stock","160.00 SAR","PRO-V","new","","PRO-V326U","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
    ["PRO-V04BU","PRO-V04BU Portable 4-inch Label Printer (Bluetooth)","Portable 4-inch thermal label printer with USB and Bluetooth.","https://nahamtech.org/product/prov-04bu","https://img.nahamtech.org/products/cashier/PRO-V04BU.jpg","in_stock","550.00 SAR","PRO-V","new","","PRO-V04BU","yes","Electronics > Print & Scan > Printers","Printers > Mobile Printers"],
    ["PRO-V300BU","PRO-V300BU Portable 3-inch Receipt Printer (Bluetooth)","Portable 3-inch thermal receipt printer with USB and Bluetooth.","https://nahamtech.org/product/prov-300bu","https://img.nahamtech.org/products/cashier/PRO-V300BU.jpg","in_stock","350.00 SAR","PRO-V","new","","PRO-V300BU","yes","Electronics > Print & Scan > Printers","Printers > Mobile Printers"],
    ["PRO-V326USE","PRO-V326USE Thermal Receipt Printer (Network)","80mm thermal receipt printer with multi-port network connectivity.","https://nahamtech.org/product/prov-326use","https://img.nahamtech.org/products/cashier/PRO-V326USE.jpg","in_stock","180.00 SAR","PRO-V","new","","PRO-V326USE","yes","Electronics > Print & Scan > Printers","Printers > Receipt Printers"],
]


# ──────────────────────────────────────────────
# GOOGLE ADS CAMPAIGN DATA
# ──────────────────────────────────────────────
ADS_HEADERS = [
    "Ad Group", "Keyword (Exact Match)", "Match Type",
    "Headline 1", "Headline 2", "Headline 3",
    "Description Line 1", "Description Line 2"
]

ADS_ROWS = [
    # AG1 - Time Attendance General
    ["Time Attendance - General","[zkteco time attendance]","Exact","ZKTeco Attendance Devices - Official","Face & Fingerprint Recognition","Free Shipping in Saudi Arabia","Shop ZKTeco time attendance machines. Face, fingerprint & card recognition. VAT-inclusive prices. Order now from NahamTech.","Trusted by Saudi businesses. ZKTeco MB2000, UFace800, SpeedFace V5L & more. Fast delivery across KSA."],
    ["Time Attendance - General","[zkteco attendance system]","Exact","","","","",""],
    ["Time Attendance - General","[جهاز بصمة حضور وانصراف]","Exact","","","","",""],
    ["Time Attendance - General","[zkteco attendance machine saudi]","Exact","","","","",""],
    ["Time Attendance - General","[جهاز حضور وانصراف بصمة وجه]","Exact","","","","",""],
    ["Time Attendance - General","[time attendance device riyadh]","Exact","","","","",""],
    ["Time Attendance - General","[fingerprint attendance machine]","Exact","","","","",""],
    ["Time Attendance - General","[جهاز بصمة zkteco]","Exact","","","","",""],
    # AG2 - MB2000
    ["MB2000","[zkteco mb2000]","Exact","ZKTeco MB2000 - Face & Fingerprint","450 SAR | VAT Inclusive","Buy Now - Fast KSA Delivery","ZKTeco MB2000 multi-biometric terminal. Face recognition + fingerprint. Ideal for offices & factories. Order online with free support.",""],
    ["MB2000","[mb2000 attendance device]","Exact","","","","",""],
    ["MB2000","[zkteco mb2000 price saudi]","Exact","","","","",""],
    ["MB2000","[zkteco mb2000 سعر]","Exact","","","","",""],
    ["MB2000","[جهاز بصمة mb2000]","Exact","","","","",""],
    # AG3 - UFace800
    ["UFace800","[zkteco uface800]","Exact","ZKTeco UFace800 - Touchscreen","Face + Finger + Card + PIN","From 810 SAR | Shop Now","UFace800 multi-biometric device with 8\" touchscreen. WiFi available. Supports face, fingerprint, card & password. Buy in Saudi Arabia.",""],
    ["UFace800","[uface 800 attendance]","Exact","","","","",""],
    ["UFace800","[zkteco uface800 wifi]","Exact","","","","",""],
    ["UFace800","[uface800 price saudi arabia]","Exact","","","","",""],
    ["UFace800","[جهاز بصمة وجه uface800]","Exact","","","","",""],
    ["UFace800","[zkteco uface800 سعر]","Exact","","","","",""],
    # AG4 - SpeedFace V5L
    ["SpeedFace V5L WiFi","[zkteco speedface v5l]","Exact","SpeedFace V5L WiFi - Palm & Face","950 SAR | Advanced Biometric","ZKTeco Authorized Distributor","ZKTeco SpeedFace V5L with WiFi. Face, palm & PIN recognition. Contactless and hygienic. Ships across Saudi Arabia.",""],
    ["SpeedFace V5L WiFi","[speedface v5l wifi]","Exact","","","","",""],
    ["SpeedFace V5L WiFi","[zkteco speedface v5l wifi price]","Exact","","","","",""],
    ["SpeedFace V5L WiFi","[جهاز بصمة كف اليد zkteco]","Exact","","","","",""],
    ["SpeedFace V5L WiFi","[palm recognition attendance device]","Exact","","","","",""],
    ["SpeedFace V5L WiFi","[speedface v5l سعر]","Exact","","","","",""],
    # AG5 - MB10
    ["MB10","[zkteco mb10]","Exact","ZKTeco MB10 - Face & Fingerprint","Only 250 SAR | Budget Biometric","Buy Online - KSA Delivery","ZKTeco MB10 compact multi-biometric terminal. Face + fingerprint verification. Best value attendance solution for small businesses.",""],
    ["MB10","[mb10 attendance machine]","Exact","","","","",""],
    ["MB10","[zkteco mb10 price]","Exact","","","","",""],
    ["MB10","[جهاز بصمة وجه mb10]","Exact","","","","",""],
    ["MB10","[zkteco mb10 سعر]","Exact","","","","",""],
    # AG6 - F18
    ["F18 Access Control","[zkteco f18]","Exact","ZKTeco F18 - Access Control","Fingerprint + Card + PIN","450 SAR | IP65 Rated","ZKTeco F18 biometric access control terminal. Fingerprint, RFID card & password. Ideal for doors and gates. Available in KSA.",""],
    ["F18 Access Control","[f18 access control]","Exact","","","","",""],
    ["F18 Access Control","[zkteco f18 fingerprint reader]","Exact","","","","",""],
    ["F18 Access Control","[جهاز تحكم في الدخول f18]","Exact","","","","",""],
    ["F18 Access Control","[zkteco f18 سعر]","Exact","","","","",""],
    ["F18 Access Control","[fingerprint access control saudi]","Exact","","","","",""],
    # AG7 - F22
    ["F22 Access Control WiFi","[zkteco f22]","Exact","ZKTeco F22 WiFi - Access Control","Finger + Card + PIN | 330 SAR","Wireless Setup - Easy Install","ZKTeco F22 with WiFi. Biometric access control with fingerprint, RFID & PIN. Wireless connectivity for easy installation. Ships in KSA.",""],
    ["F22 Access Control WiFi","[zkteco f22 wifi]","Exact","","","","",""],
    ["F22 Access Control WiFi","[f22 access control wifi]","Exact","","","","",""],
    ["F22 Access Control WiFi","[جهاز تحكم في الدخول f22]","Exact","","","","",""],
    ["F22 Access Control WiFi","[zkteco f22 سعر]","Exact","","","","",""],
    # AG8 - BioTime
    ["BioTime Software","[zkteco biotime]","Exact","ZKTeco BioTime - Web Software","Cloud-Based Attendance System","Setup & Consultation Available","BioTime by ZKTeco - powerful web-based attendance management software. Professional setup and consultation. Contact NahamTech today.",""],
    ["BioTime Software","[biotime software]","Exact","","","","",""],
    ["BioTime Software","[برنامج حضور وانصراف zkteco]","Exact","","","","",""],
    ["BioTime Software","[zkteco biotime download]","Exact","","","","",""],
    ["BioTime Software","[biotime attendance software]","Exact","","","","",""],
    ["BioTime Software","[biotime consultation]","Exact","","","","",""],
    # AG9 - Access Control General
    ["Access Control Systems","[zkteco access control]","Exact","ZKTeco Access Control Systems","Locks, Readers & Controllers","Professional Installation in KSA","Complete ZKTeco access control solutions. Magnetic locks, smart glass locks, fingerprint readers & exit buttons. Serving all of Saudi Arabia.",""],
    ["Access Control Systems","[أنظمة التحكم في الدخول]","Exact","","","","",""],
    ["Access Control Systems","[magnetic lock zkteco]","Exact","","","","",""],
    ["Access Control Systems","[zkteco door lock saudi]","Exact","","","","",""],
    ["Access Control Systems","[قفل مغناطيسي zkteco]","Exact","","","","",""],
    ["Access Control Systems","[zkteco glass door lock]","Exact","","","","",""],
    # AG10 - USB Readers
    ["USB Fingerprint Readers","[zkteco zk4500]","Exact","USB Fingerprint Readers - ZKTeco","ZK4500, ZK9500 & SLK20R","Plug & Play | SDK Included","ZKTeco USB fingerprint scanners. ZK4500, ZK9500 & SLK20R. Compatible with BioTime & third-party HR software. Ships across KSA.",""],
    ["USB Fingerprint Readers","[zkteco zk9500]","Exact","","","","",""],
    ["USB Fingerprint Readers","[zkteco slk20r]","Exact","","","","",""],
    ["USB Fingerprint Readers","[usb fingerprint reader saudi]","Exact","","","","",""],
    ["USB Fingerprint Readers","[قارئ بصمة usb]","Exact","","","","",""],
    ["USB Fingerprint Readers","[fingerprint scanner usb]","Exact","","","","",""],
    # AG11 - Smart Locks / Biosecurity
    ["Smart Locks & Biosecurity","[zkteco smart lock]","Exact","ZKTeco Smart Locks & Security","Glass Door Lock GL300 | From 550 SAR","Touchless Exit Buttons Available","ZKTeco biosecurity products. GL300 smart glass door lock, TLEB101 touchless exit button, magnetic locks & mounting brackets. Shop now.",""],
    ["Smart Locks & Biosecurity","[zkteco gl300]","Exact","","","","",""],
    ["Smart Locks & Biosecurity","[قفل ذكي زجاج zkteco]","Exact","","","","",""],
    ["Smart Locks & Biosecurity","[smart glass door lock saudi]","Exact","","","","",""],
    ["Smart Locks & Biosecurity","[zkteco biosecurity]","Exact","","","","",""],
    ["Smart Locks & Biosecurity","[zkteco exit button]","Exact","","","","",""],
]


# ═══════════════════════════════════════════════
#  1. GMC FEED - XLSX
# ═══════════════════════════════════════════════
print("[1/6] Writing GMC Feed XLSX...")
wb = Workbook()
ws = wb.active
ws.title = "GMC Product Feed"
ws.sheet_properties.tabColor = "2F5496"

ws.append(GMC_HEADERS)
for row in GMC_ROWS:
    ws.append(row)

style_header(ws, len(GMC_HEADERS))
style_data(ws, len(GMC_ROWS) + 1, len(GMC_HEADERS))

# Conditional formatting for availability
for row_idx in range(2, len(GMC_ROWS) + 2):
    avail_cell = ws.cell(row=row_idx, column=6)
    if avail_cell.value == "in_stock":
        avail_cell.fill = GREEN_FILL
    elif avail_cell.value == "out_of_stock":
        avail_cell.fill = RED_FILL

auto_width(ws, len(GMC_HEADERS))
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(f"{OUT_DIR}/gmc_product_feed.xlsx")
print("   -> gmc_product_feed.xlsx")


# ═══════════════════════════════════════════════
#  2. GMC FEED - CSV (UTF-8 BOM for Excel compat)
# ═══════════════════════════════════════════════
print("[2/6] Writing GMC Feed CSV...")
with open(f"{OUT_DIR}/gmc_product_feed.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(GMC_HEADERS)
    writer.writerows(GMC_ROWS)
print("   -> gmc_product_feed.csv")


# ═══════════════════════════════════════════════
#  3. GMC FEED - JSON
# ═══════════════════════════════════════════════
print("[3/6] Writing GMC Feed JSON...")
feed_json = []
for row in GMC_ROWS:
    feed_json.append(dict(zip(GMC_HEADERS, row)))
with open(f"{OUT_DIR}/gmc_product_feed.json", "w", encoding="utf-8") as f:
    json.dump(feed_json, f, indent=2, ensure_ascii=False)
print("   -> gmc_product_feed.json")


# ═══════════════════════════════════════════════
#  4. GOOGLE ADS CAMPAIGN - XLSX
# ═══════════════════════════════════════════════
print("[4/6] Writing Google Ads Campaign XLSX...")
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Google Ads Campaign"
ws2.sheet_properties.tabColor = "C00000"

ws2.append(ADS_HEADERS)
for row in ADS_ROWS:
    ws2.append(row)

style_header(ws2, len(ADS_HEADERS))
style_data(ws2, len(ADS_ROWS) + 1, len(ADS_HEADERS))

# Color-code ad groups
AG_COLORS = {
    "Time Attendance - General": "DAEEF3",
    "MB2000": "FDE9D9",
    "UFace800": "E4DFEC",
    "SpeedFace V5L WiFi": "D8E4BC",
    "MB10": "F2DCDB",
    "F18 Access Control": "DCE6F1",
    "F22 Access Control WiFi": "EBD1D8",
    "BioTime Software": "FFFFCC",
    "Access Control Systems": "D9E9D9",
    "USB Fingerprint Readers": "FDEADA",
    "Smart Locks & Biosecurity": "E6E6E6",
}
for row_idx in range(2, len(ADS_ROWS) + 2):
    ag_name = ws2.cell(row=row_idx, column=1).value
    if ag_name in AG_COLORS:
        fill = PatternFill(start_color=AG_COLORS[ag_name], end_color=AG_COLORS[ag_name], fill_type="solid")
        for col_idx in range(1, len(ADS_HEADERS) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = fill

auto_width(ws2, len(ADS_HEADERS))
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

wb2.save(f"{OUT_DIR}/google_ads_campaign.xlsx")
print("   -> google_ads_campaign.xlsx")


# ═══════════════════════════════════════════════
#  5. GOOGLE ADS CAMPAIGN - CSV
# ═══════════════════════════════════════════════
print("[5/6] Writing Google Ads Campaign CSV...")
with open(f"{OUT_DIR}/google_ads_campaign.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(ADS_HEADERS)
    writer.writerows(ADS_ROWS)
print("   -> google_ads_campaign.csv")


# ═══════════════════════════════════════════════
#  6. GOOGLE ADS CAMPAIGN - JSON
# ═══════════════════════════════════════════════
print("[6/6] Writing Google Ads Campaign JSON...")
ads_json = []
for row in ADS_ROWS:
    ads_json.append(dict(zip(ADS_HEADERS, row)))
with open(f"{OUT_DIR}/google_ads_campaign.json", "w", encoding="utf-8") as f:
    json.dump(ads_json, f, indent=2, ensure_ascii=False)
print("   -> google_ads_campaign.json")


# ═══════════════════════════════════════════════
#  SUMMARY
# ═══════════════════════════════════════════════
print("\n" + "=" * 55)
print("  ALL FILES EXPORTED SUCCESSFULLY")
print("=" * 55)
print(f"  Output directory: {OUT_DIR}/")
print(f"  GMC Feed:        {len(GMC_ROWS)} products")
print(f"  Ad Groups:       {len(AG_COLORS)} groups, {len(ADS_ROWS)} keyword rows")
print("=" * 55)
print("\nFiles:")
for fname in sorted(os.listdir(OUT_DIR)):
    fpath = os.path.join(OUT_DIR, fname)
    size_kb = os.path.getsize(fpath) / 1024
    print(f"  {fname:40s} {size_kb:>7.1f} KB")
